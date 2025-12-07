from openpyxl import load_workbook

x5_professinon = [
    "53830373",
    "53830677",
    "53830682",
    "53830687",
    "53830691",
    "53830374",
    "53830679",
    "53830684",
    "53830688",
    "53830692",
    "53830372",
    "53830676",
    "53830681",
    "53830686",
    "53830690",
    "53830675",
    "53830680",
    "53830685",
    "53830689",
    "53830693"
]

x5_professinon_rus = [
    "Грузчик СМЗ БТС ТС5",
    "Грузчик СМЗ К1 ТС5",
    "Грузчик СМЗ К2 ТС5",
    "Грузчик СМЗ К3 ТС5",
    "Грузчик СМЗ К4 ТС5",
    "Пекарь СМЗ БТС ТС5",
    "Пекарь СМЗ К1 ТС5",
    "Пекарь СМЗ К2 ТС5",
    "Пекарь СМЗ К3 ТС5",
    "Пекарь СМЗ К4 ТС5",
    "Продавец СМЗ БТС ТС5",
    "Продавец СМЗ К1 ТС5",
    "Продавец СМЗ К2 ТС5",
    "Продавец СМЗ К3 ТС5",
    "Продавец СМЗ К4 ТС5",
    "Сборщик заказов СМЗ БТС ТС5",
    "Сборщик заказов СМЗ К1 ТС5",
    "Сборщик заказов СМЗ К2 ТС5",
    "Сборщик заказов СМЗ К3 ТС5",
    "Сборщик заказов СМЗ К4 ТС5"
]


class TT:
    def __init__(self, wop_id, tarif_zone, group_id):
        self.wop_id = wop_id
        self.tarif_zone = tarif_zone
        self.group_id = group_id
        self.tariff_object = None

class Tariff:
    def __init__(self, region, gruzchik=None, prodavec=None, sbopshik=None, pekar=None):
        self.region = region
        self.professions = {
            'gruzchik': gruzchik,
            'prodavec': prodavec,
            'sbopshik': sbopshik,
            'pekar': pekar
        }


def read_tt_file(sheet):
    result_dict = []
    line = 2
    for row in sheet.iter_rows(min_row=2):
        if any(cell.value is not None for cell in row):

            wop_id = sheet[f'F{line}'].value # Ячейка Wop_id
            tarif = sheet[f'S{line}'].value
            group_id = sheet[f'N{line}'].value

            if wop_id != None:
                wop_id = wop_id.split("_")[-1]
            else:
                TypeError("Неверный тип")

            tt_x5 = TT(wop_id, tarif, group_id)
            result_dict.append(tt_x5)
            line += 1

    return result_dict


def load_tarif_from_file(sheet, zone):
    proff_dict = {}
    line = 2
    for row in sheet.iter_rows(min_row=2):
        if any(cell.value is not None for cell in row):
            if sheet[f'B{line}'].value == zone:
                proff_dict[sheet[f'C{line}'].value] = [
                    round(sheet[f'E{line}'].value, 2),
                    round(sheet[f'F{line}'].value, 2),
                    round(sheet[f'G{line}'].value, 2),
                    round(sheet[f'H{line}'].value, 2),
                    round(sheet[f'I{line}'].value, 2)
                ]
            line += 1


    tariff = Tariff(
        region=zone,
        gruzchik=proff_dict['Услуги по погрузке-разгрузке товара'],
        sbopshik=proff_dict['Услуга по сборке товара в торговом зале'],
        prodavec=proff_dict['Услуги по выкладке и предпродажной подготовке товара'],
        pekar=proff_dict['Услуги по формовке и выпечке хлебобулочных изделий']
    )

    return tariff



def create_sample(template_sheet, wop_id, group_id, tarif_data, start):
    """
    Заполняем итоговый фаил.
    :param template_sheet:
    :param tarif_data:
    :param wop_id:
    :param group_id:
    :param start:
    :return:
    """
    current_row = start

    for i in range(0, 20, 5):
        for j in range(5):
            value_prof = x5_professinon[i+j]
            name_field = x5_professinon_rus[i+j]

            prof_id = "x5_" + value_prof
            id_field = "payment_x5_" + wop_id + "_" + value_prof

            if name_field[:-8] == "Грузчик СМЗ":
                tariffs = tarif_data.professions["gruzchik"]
            elif name_field[:-8] == "Пекарь СМЗ":
                tariffs = tarif_data.professions["pekar"]
            elif name_field[:-8] == "Продавец СМЗ":
                tariffs = tarif_data.professions["prodavec"]
            elif name_field[:-8] == "Сборщик заказов СМЗ":
                tariffs = tarif_data.professions["sbopshik"]

            template_sheet.cell(row=current_row + i + j, column=1, value=id_field)
            template_sheet.cell(row=current_row + i + j, column=3, value=name_field)
            template_sheet.cell(row=current_row + i + j, column=12, value=prof_id)


            template_sheet.cell(row=current_row + i + j, column=6, value=tariffs[j])

    current_row += 20

    return current_row





def main():
    # Загружаем данные ТТ
    wb_source = load_workbook('Новые ТТ.xlsx')
    source_sheet = wb_source.active
    print("Чтение файла с точками завершено")

    # Читаем данные
    data_dict = read_tt_file(source_sheet)

    #  Загружаем данные Тарифов
    tarif_source = load_workbook("Тарифы Пятёрочка.xlsx", data_only=True)
    source_sheet_tarif = tarif_source.active


    # Загружаем шаблон ДЛЯ ЗАПИСИ
    wb_template = load_workbook('payments_x5.xlsx')
    template_sheet = wb_template.active

    count = 0

    current_position = 2 # С какой строки нужно заполнять таблицу
    for tt in data_dict:
        count += 1
        print(tt.wop_id, "регион", tt.tarif_zone)
        tarif_data = load_tarif_from_file(source_sheet_tarif, tt.tarif_zone)

        current_position = create_sample(template_sheet, tt.wop_id, tt.group_id, tarif_data, current_position)

    wb_template.save('payments_x5.xlsx')
    print("Файл успешно сохранен!")
    print(f"Количество точек {count}")
    input("Нажмите любую клавишу")




    # Закрываем файлы
    tarif_source.close()
    wb_template.close()
    wb_source.close()

if __name__ == "__main__":
    input("Убедись что в файле с точками в поле S нет сокращений обл. Нужно что бы было написано область")
    main()
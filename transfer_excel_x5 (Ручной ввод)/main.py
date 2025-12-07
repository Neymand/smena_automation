from openpyxl import load_workbook

x5_professinon = [
    "53830373",
    "53830677",
    "53830682",
    "53830687",
    "53830691",
    "53830374",
    "53830679",
    "53830684",
    "53830688",
    "53830692",
    "53830372",
    "53830676",
    "53830681",
    "53830686",
    "53830690",
    "53830675",
    "53830680",
    "53830685",
    "53830689",
    "53830693"
]

x5_professinon_rus = [
    "Грузчик СМЗ БТС ТС5",
    "Грузчик СМЗ К1 ТС5",
    "Грузчик СМЗ К2 ТС5",
    "Грузчик СМЗ К3 ТС5",
    "Грузчик СМЗ К4 ТС5",
    "Пекарь СМЗ БТС ТС5",
    "Пекарь СМЗ К1 ТС5",
    "Пекарь СМЗ К2 ТС5",
    "Пекарь СМЗ К3 ТС5",
    "Пекарь СМЗ К4 ТС5",
    "Продавец СМЗ БТС ТС5",
    "Продавец СМЗ К1 ТС5",
    "Продавец СМЗ К2 ТС5",
    "Продавец СМЗ К3 ТС5",
    "Продавец СМЗ К4 ТС5",
    "Сборщик заказов СМЗ БТС ТС5",
    "Сборщик заказов СМЗ К1 ТС5",
    "Сборщик заказов СМЗ К2 ТС5",
    "Сборщик заказов СМЗ К3 ТС5",
    "Сборщик заказов СМЗ К4 ТС5"
]


class TT:
    def __init__(self, wop_id, tarif_zone):
        self.wop_id = wop_id
        self.tarif_zone = tarif_zone
        #self.group_id = group_id
        self.tariff_object = None

class Tariff:
    def __init__(self, region, gruzchik=None, prodavec=None, sbopshik=None, pekar=None):
        self.region = region
        self.professions = {
            'gruzchik': gruzchik,
            'prodavec': prodavec,
            'sbopshik': sbopshik,
            'pekar': pekar
        }

def find_first_empty_row(sheet):
    """Находит первую полностью пустую строку в листе"""
    row = 1
    while True:
        if all(cell.value is None for cell in sheet[row]):
            return row
        row += 1

def manual_input_points():
    wop_id = input("Введите Wop_id точки: ")
    region = input("Введите регион: ")

    retail_point = TT(wop_id, region)
    return retail_point

def load_tarif_from_file(sheet, zone):
    proff_dict = {}
    line = 2
    for row in sheet.iter_rows(min_row=2):
        if any(cell.value is not None for cell in row):
            if sheet[f'B{line}'].value == zone:
                proff_dict[sheet[f'C{line}'].value] = [
                    int(round(sheet[f'E{line}'].value, 2)) * 100,
                    int(round(sheet[f'F{line}'].value, 2)) * 100,
                    int(round(sheet[f'G{line}'].value, 2)) * 100,
                    int(round(sheet[f'H{line}'].value, 2)) * 100,
                    int(round(sheet[f'I{line}'].value, 2)) * 100
                ]
            line += 1

    tariff = Tariff(
        region=zone,
        gruzchik=proff_dict['Услуги по погрузке-разгрузке товара'],
        sbopshik=proff_dict['Услуга по сборке товара в торговом зале'],
        prodavec=proff_dict['Услуги по выкладке и предпродажной подготовке товара'],
        pekar=proff_dict['Услуги по формовке и выпечке хлебобулочных изделий']
    )

    return tariff

def create_sample(template_sheet, wop_id, tarif_data):
    """
    Заполняем итоговый фаил.
    :param template_sheet:
    :param tarif_data:
    :param wop_id:
    :param group_id:
    :param start:
    :return:
    """
    current_row = find_first_empty_row(template_sheet)

    for i in range(0, 20, 5):
        for j in range(5):
            value_prof = x5_professinon[i+j]
            name_field = x5_professinon_rus[i+j]

            prof_id = "x5_" + value_prof
            id_field = "payment_x5_" + wop_id + "_" + value_prof

            if name_field[:-8] == "Грузчик СМЗ":
                tariffs = tarif_data.professions["gruzchik"]
            elif name_field[:-8] == "Пекарь СМЗ":
                tariffs = tarif_data.professions["pekar"]
            elif name_field[:-8] == "Продавец СМЗ":
                tariffs = tarif_data.professions["prodavec"]
            elif name_field[:-8] == "Сборщик заказов СМЗ":
                tariffs = tarif_data.professions["sbopshik"]

            template_sheet.cell(row=current_row + i + j, column=1, value=id_field)
            template_sheet.cell(row=current_row + i + j, column=3, value=name_field)
            template_sheet.cell(row=current_row + i + j, column=12, value=prof_id)


            template_sheet.cell(row=current_row + i + j, column=6, value=tariffs[j])

    current_row += 20

    return current_row


def main():


    tt = manual_input_points() # Объект ТТ

    '''Фаил с тарифами'''
    tarif_source = load_workbook("Тарифы Пятёрочка.xlsx", data_only=True)
    source_sheet_tarif = tarif_source.active

    '''Фаил где будет результат'''
    wb_template = load_workbook('payments_x5.xlsx')
    template_sheet = wb_template.active

    """Заполняем фаил"""
    tarif_data = load_tarif_from_file(source_sheet_tarif, tt.tarif_zone)
    create_sample(template_sheet, tt.wop_id, tarif_data)



    wb_template.save('payments_x5.xlsx')
    print("Файл успешно сохранен!")



    # Закрываем файлы
    tarif_source.close()
    wb_template.close()


if __name__ == "__main__":
    while True:
        main()

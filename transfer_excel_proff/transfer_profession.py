from openpyxl import load_workbook




def read_excel_to_dict(sheet):
    result_dict = {}
    # Определяем количество столбцов
    max_column = sheet.max_column

    # Проходим по всем столбцам первой строки
    for col in range(1, max_column + 1):
        header_cell = sheet.cell(row=1, column=col)
        header_value = header_cell.value
        header_name = str(header_value)

        # Получаем все значения под этим заголовком в столбце
        column_values = []
        for row in range(2, sheet.max_row + 1):  # Начинаем со второй строки
            cell = sheet.cell(row=row, column=col)

            # Обрабатываем пустые ячейки
            if cell.value is None:
                column_values.append(None)
            elif isinstance(cell.value, str) and cell.value.strip() == '':
                column_values.append(None)
            else:
                column_values.append(cell.value)

        # Добавляем в словарь
        result_dict[header_name] = column_values
    return result_dict




def add_in_templexcel(template_sheet, data_dict):
    cells_modified = 0
    for row in template_sheet.iter_rows():
        for cell in row:
            # Проверяем, есть ли значение ячейки в ключах словаря
            if cell.value and str(cell.value) in data_dict:
                header_name = str(cell.value)
                values = data_dict[header_name]


                for i, value in enumerate(values, start=1):
                    target_row = cell.row + i
                    target_cell = template_sheet.cell(row=target_row, column=cell.column)
                    if value is None:
                        target_cell.value = None
                    else:
                        target_cell.value = value
                    cells_modified += 1
        print(f"Изменено ячеек: {cells_modified}")
        return cells_modified

# Использование
def main():
    # Загружаем исходные данные
    wb_source = load_workbook('professions_blue.xlsx')
    source_sheet = wb_source.active

    # Читаем данные в словарь (нужно реализовать эту функцию)
    data_dict = read_excel_to_dict(source_sheet)

    # Загружаем шаблон ДЛЯ ЗАПИСИ
    wb_template = load_workbook('шаблон профессий.xlsx')
    template_sheet = wb_template.active

    # Записываем данные в шаблон
    add_in_templexcel(template_sheet, data_dict)

    # Сохраняем изменения в файл!
    wb_template.save('шаблон профессий.xlsx')
    print("Файл успешно сохранен!")

    # Закрываем файлы
    wb_source.close()
    wb_template.close()

if __name__ == "__main__":
    main()






